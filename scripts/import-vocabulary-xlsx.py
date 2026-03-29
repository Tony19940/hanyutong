import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


WORD_PATTERN = re.compile(r"^\s*(?P<chinese>.+?)\s*\((?P<pinyin>.+?)\)\s*$")
EXAMPLE_PATTERN = re.compile(
    r"^\s*(?P<cn>.+?)\s*\((?P<pinyin>.+?)\)\s*(?P<km>.+?)\s*$"
)


def parse_word(raw_value):
    if raw_value is None:
        return "", ""

    text = str(raw_value).strip()
    match = WORD_PATTERN.match(text)
    if not match:
        return text, ""
    return match.group("chinese").strip(), match.group("pinyin").strip()


def parse_example(raw_value):
    if raw_value is None:
        return None

    text = str(raw_value).strip()
    if not text:
        return None

    match = EXAMPLE_PATTERN.match(text)
    if match:
        return {
            "chinese": match.group("cn").strip(),
            "pinyin": match.group("pinyin").strip(),
            "khmer": match.group("km").strip(),
            "audio": None,
        }

    return {
        "chinese": text,
        "pinyin": "",
        "khmer": "",
        "audio": None,
    }


def build_entry(row, entry_id):
    sequence, word_cell, khmer_cell, example_1_cell, example_2_cell = row[:5]
    chinese, pinyin = parse_word(word_cell)
    examples = [parse_example(example_1_cell), parse_example(example_2_cell)]
    examples = [example for example in examples if example]

    primary = examples[0] if examples else None
    secondary = examples[1] if len(examples) > 1 else None

    return {
        "id": entry_id,
        "source_number": int(float(sequence)),
        "chinese": chinese,
        "pinyin": pinyin,
        "khmer": str(khmer_cell).strip() if khmer_cell else "",
        "emoji": None,
        "visual": None,
        "category": "core",
        "hsk": 1,
        "examples": examples,
        "example_cn": primary["chinese"] if primary else "",
        "example_pinyin": primary["pinyin"] if primary else "",
        "example_km": primary["khmer"] if primary else "",
        "audio_word": None,
        "audio_example": primary["audio"] if primary else None,
        "example_cn_2": secondary["chinese"] if secondary else "",
        "example_pinyin_2": secondary["pinyin"] if secondary else "",
        "example_km_2": secondary["khmer"] if secondary else "",
        "audio_example_2": secondary["audio"] if secondary else None,
    }


def main():
    if len(sys.argv) < 3:
        print(
            "Usage: python scripts/import-vocabulary-xlsx.py <input.xlsx> <output.json>",
            file=sys.stderr,
        )
        sys.exit(1)

    source = Path(sys.argv[1]).expanduser().resolve()
    target = Path(sys.argv[2]).expanduser().resolve()

    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    vocabulary = []
    next_id = 1
    for row in rows:
        if not row or not row[0]:
            continue
        try:
            float(row[0])
        except (TypeError, ValueError):
            continue
        vocabulary.append(build_entry(row, next_id))
        next_id += 1

    target.write_text(
        json.dumps(vocabulary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Imported {len(vocabulary)} entries from {source} -> {target}")


if __name__ == "__main__":
    main()
